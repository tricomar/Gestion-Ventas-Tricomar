"""
Rutas para importación y exportación de datos
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from utils.auth import get_current_user, User
from utils.database import db
from middleware.tenant import get_tenant_filter
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime
from uuid import uuid4

router = APIRouter()

@router.get("/products/template")
async def download_products_template(current_user: User = Depends(get_current_user)):
    """
    Descarga una plantilla Excel con el formato para importar productos
    Solo para account_admin y supervisor
    """
    # Verificar permisos
    if current_user.role not in ['account_admin', 'supervisor', 'admin']:
        raise HTTPException(
            status_code=403,
            detail="No tienes permisos para descargar plantillas"
        )
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Encabezados
    headers = [
        "Nombre del Producto",
        "Código de Tienda",
        "Precio de Compra",
        "Precio de Venta",
        "Stock Disponible",
        "Categoría (opcional)"
    ]
    ws.append(headers)
    
    # Estilo para encabezados
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Agregar ejemplos
    ws.append([
        "Cat Chow Pescado 15kg",
        "PT",
        2400,
        3500,
        10,
        "Alimentos para Mascotas"
    ])
    ws.append([
        "Pipa de Vidrio",
        "ST",
        520,
        5000,
        5,
        "Accesorios"
    ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    
    # Agregar instrucciones en una hoja separada
    ws_instrucciones = wb.create_sheet("Instrucciones")
    ws_instrucciones.append(["INSTRUCCIONES PARA IMPORTAR PRODUCTOS"])
    ws_instrucciones.append([])
    ws_instrucciones.append(["1. Complete la información de cada producto en la hoja 'Productos'"])
    ws_instrucciones.append(["2. Campos obligatorios: Nombre, Código de Tienda, Precio de Compra, Precio de Venta"])
    ws_instrucciones.append(["3. El Código de Tienda debe coincidir con los códigos de sus tiendas (ej: PT, ST, TT)"])
    ws_instrucciones.append(["4. Los precios deben ser números positivos"])
    ws_instrucciones.append(["5. El Stock Disponible es opcional, por defecto será 0"])
    ws_instrucciones.append(["6. Guarde el archivo y súbalo desde Configuración > Importar Datos"])
    ws_instrucciones.append([])
    ws_instrucciones.append(["NOTA: Los productos duplicados (mismo nombre y tienda) serán actualizados"])
    
    ws_instrucciones.column_dimensions['A'].width = 80
    ws_instrucciones[1][0].font = Font(bold=True, size=14)
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retornar como descarga
    filename = f"plantilla_productos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/products/import")
async def import_products(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Importa productos desde un archivo Excel
    Solo para account_admin y supervisor
    """
    # Verificar permisos
    if current_user.role not in ['account_admin', 'supervisor', 'admin']:
        raise HTTPException(
            status_code=403,
            detail="No tienes permisos para importar productos"
        )
    
    # Verificar que sea un archivo Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    try:
        # Leer archivo
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Verificar encabezados
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Nombre del Producto",
            "Código de Tienda",
            "Precio de Compra",
            "Precio de Venta",
            "Stock Disponible",
            "Categoría (opcional)"
        ]
        
        if headers[:4] != expected_headers[:4]:
            raise HTTPException(
                status_code=400,
                detail="El formato del archivo no es válido. Use la plantilla proporcionada."
            )
        
        # Obtener las tiendas de la cuenta
        account = await db.accounts.find_one({"id": current_user.account_id}, {"_id": 0})
        if not account:
            raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        
        store_codes = {store.get('code'): store.get('id') for store in account.get('stores', [])}
        
        # Procesar filas
        products_added = 0
        products_updated = 0
        errors = []
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Fila vacía
                continue
            
            try:
                name = row[0]
                store_code = row[1]
                cost_price = row[2]
                sale_price = row[3]
                stock = row[4] if len(row) > 4 and row[4] is not None else 0
                category = row[5] if len(row) > 5 else ""
                
                # Validaciones
                if not name or not store_code or cost_price is None or sale_price is None:
                    errors.append(f"Fila {idx}: Faltan campos obligatorios")
                    continue
                
                if store_code not in store_codes:
                    errors.append(f"Fila {idx}: Código de tienda '{store_code}' no válido")
                    continue
                
                if cost_price < 0 or sale_price < 0:
                    errors.append(f"Fila {idx}: Los precios deben ser positivos")
                    continue
                
                # Buscar si ya existe el producto
                existing_product = await db.products.find_one({
                    "account_id": current_user.account_id,
                    "name": name,
                    "store": store_code
                }, {"_id": 0})
                
                product_data = {
                    "account_id": current_user.account_id,
                    "name": str(name).strip(),
                    "store": store_code,
                    "cost_price": float(cost_price),
                    "sale_price": float(sale_price),
                    "stock": int(stock) if stock else 0,
                    "category": str(category).strip() if category else "",
                    "usage_count": existing_product.get("usage_count", 0) if existing_product else 0,
                    "created_at": existing_product.get("created_at") if existing_product else datetime.now().isoformat()
                }
                
                if existing_product:
                    # Actualizar producto existente
                    await db.products.update_one(
                        {"id": existing_product["id"]},
                        {"$set": product_data}
                    )
                    products_updated += 1
                else:
                    # Crear nuevo producto
                    product_data["id"] = str(uuid4())
                    await db.products.insert_one(product_data)
                    products_added += 1
                
            except Exception as e:
                errors.append(f"Fila {idx}: {str(e)}")
        
        return {
            "success": True,
            "products_added": products_added,
            "products_updated": products_updated,
            "errors": errors,
            "message": f"Se procesaron {products_added + products_updated} productos correctamente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al procesar el archivo: {str(e)}"
        )
